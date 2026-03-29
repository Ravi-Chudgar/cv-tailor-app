"""
User Storage Management using Excel
Handles persistent user data storage in Excel file
"""

import os
import pandas as pd
from datetime import datetime
from typing import Optional, Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import bcrypt

# Excel file path
USERS_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'users.xlsx')

# Create data directory if it doesn't exist
os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)


def hash_password(password: str) -> str:
    """Hash a password using bcrypt"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(password: str, hashed: str) -> bool:
    """Verify a password against its hash"""
    try:
        return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        return False


def ensure_users_file():
    """Create users.xlsx if it doesn't exist with admin user"""
    if not os.path.exists(USERS_FILE):
        # Create workbook with headers
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Define headers
        headers = ['User ID', 'Username', 'Password Hash', 'Email', 'Is Active', 
                   'Created At', 'Updated At', 'Role']
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 12
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
        
        # Add default admin user
        admin_password_hash = hash_password("admin123")
        admin_row = [
            "1",
            "admin",
            admin_password_hash,
            "admin@example.com",
            True,
            datetime.now().isoformat(),
            datetime.now().isoformat(),
            "admin"
        ]
        
        for col_num, value in enumerate(admin_row, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        wb.save(USERS_FILE)
        print(f"✅ Users file created: {USERS_FILE}")


def get_all_users() -> List[Dict]:
    """Get all users from Excel file"""
    ensure_users_file()
    try:
        df = pd.read_excel(USERS_FILE, sheet_name='Users')
        users = []
        for _, row in df.iterrows():
            users.append({
                'user_id': str(row['User ID']),
                'username': row['Username'],
                'password_hash': row['Password Hash'],
                'email': row['Email'],
                'is_active': bool(row['Is Active']),
                'created_at': row['Created At'],
                'updated_at': row['Updated At'],
                'role': row['Role']
            })
        return users
    except Exception as e:
        print(f"❌ Error reading users: {e}")
        return []


def get_user_by_username(username: str) -> Optional[Dict]:
    """Get specific user by username"""
    users = get_all_users()
    for user in users:
        if user['username'].lower() == username.lower():
            return user
    return None


def get_user_by_id(user_id: str) -> Optional[Dict]:
    """Get specific user by ID"""
    users = get_all_users()
    for user in users:
        if user['user_id'] == user_id:
            return user
    return None


def add_user(username: str, password: str, email: str, role: str = "user") -> Dict:
    """Add a new user to Excel file"""
    ensure_users_file()
    
    # Check if user already exists
    if get_user_by_username(username):
        raise ValueError("Username already exists")
    
    try:
        df = pd.read_excel(USERS_FILE, sheet_name='Users')
        
        # Generate new user ID
        next_user_id = int(df['User ID'].max()) + 1
        
        # Create new user record
        password_hash = hash_password(password)
        now = datetime.now().isoformat()
        
        new_user = {
            'User ID': next_user_id,
            'Username': username,
            'Password Hash': password_hash,
            'Email': email,
            'Is Active': True,
            'Created At': now,
            'Updated At': now,
            'Role': role
        }
        
        # Append new user
        df = pd.concat([df, pd.DataFrame([new_user])], ignore_index=True)
        
        # Write back to Excel
        with pd.ExcelWriter(USERS_FILE, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Users', index=False)
        
        # Format the Excel file
        excel_format_users()
        
        return {
            'user_id': str(next_user_id),
            'username': username,
            'email': email,
            'role': role,
            'created_at': now
        }
    except Exception as e:
        raise Exception(f"Failed to add user: {str(e)}")


def excel_format_users():
    """Format the users Excel file with styles"""
    try:
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb['Users']
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        wb.save(USERS_FILE)
    except Exception as e:
        print(f"⚠️ Warning: Could not format Excel file: {e}")


def delete_user(username: str) -> bool:
    """Delete a user from Excel file"""
    if username.lower() == "admin":
        raise ValueError("Cannot delete admin user")
    
    ensure_users_file()
    try:
        df = pd.read_excel(USERS_FILE, sheet_name='Users')
        
        # Filter out the user
        df = df[df['Username'].str.lower() != username.lower()]
        
        # Write back to Excel
        with pd.ExcelWriter(USERS_FILE, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Users', index=False)
        
        # Format the Excel file
        excel_format_users()
        
        return True
    except Exception as e:
        raise Exception(f"Failed to delete user: {str(e)}")


def verify_user_credentials(username: str, password: str) -> Optional[Dict]:
    """Verify user credentials and return user if valid"""
    user = get_user_by_username(username)
    
    if not user:
        return None
    
    if not user['is_active']:
        return None
    
    if verify_password(password, user['password_hash']):
        # Don't return password hash
        user_copy = user.copy()
        del user_copy['password_hash']
        return user_copy
    
    return None


def update_user(username: str, **kwargs) -> Dict:
    """Update user information"""
    ensure_users_file()
    try:
        df = pd.read_excel(USERS_FILE, sheet_name='Users')
        
        # Find and update user
        mask = df['Username'].str.lower() == username.lower()
        if not mask.any():
            raise ValueError(f"User {username} not found")
        
        # Update allowed fields
        allowed_fields = {'Email', 'Is Active', 'Role'}
        for field, value in kwargs.items():
            if field in allowed_fields:
                df.loc[mask, field] = value
        
        df.loc[mask, 'Updated At'] = datetime.now().isoformat()
        
        # Write back to Excel
        with pd.ExcelWriter(USERS_FILE, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Users', index=False)
        
        excel_format_users()
        
        updated_user = get_user_by_username(username)
        return {k: v for k, v in updated_user.items() if k != 'password_hash'}
    except Exception as e:
        raise Exception(f"Failed to update user: {str(e)}")
